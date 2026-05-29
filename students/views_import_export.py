"""Data import/export views."""

import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook, load_workbook

from .decorators import permission_required_custom
from .models import ClassInfo, Department, Student


@permission_required_custom("students.can_manage_student")
def import_export(request):
    """Import/export page."""
    if request.method == "POST" and request.FILES.get("file"):
        return _handle_import(request)
    return render(request, "import_export.html")


def _handle_import(request):
    """Process uploaded Excel file and import students."""
    uploaded_file = request.FILES["file"]
    try:
        wb = load_workbook(uploaded_file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        messages.error(request, f"文件解析失败: {e}")
        return redirect("import_export")

    created, updated, errors = 0, 0, []
    for i, row in enumerate(rows, start=2):
        try:
            student_id, name, age, major = str(row[0]), str(row[1]), int(row[2]), str(row[3])
            enroll_date, grad_date = str(row[4]), str(row[5])
        except (IndexError, ValueError, TypeError) as e:
            errors.append(f"第{i}行: 数据格式错误 - {e}")
            continue

        obj, was_created = Student.objects.update_or_create(
            student_id=student_id,
            defaults={
                "name": name,
                "age": age,
                "major": major,
                "enrollment_date": enroll_date,
                "graduation_date": grad_date,
            },
        )
        if was_created:
            created += 1
        else:
            updated += 1

    if errors:
        for err in errors[:10]:
            messages.warning(request, err)
    messages.success(request, f"导入完成: 新增 {created}, 更新 {updated}, 失败 {len(errors)}")
    return redirect("import_export")


@permission_required_custom("students.can_export_data")
def export_students(request):
    """Export all students to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "学生列表"
    ws.append(["学号", "姓名", "年龄", "专业", "班级", "入学时间", "毕业时间"])

    for s in Student.objects.select_related("class_info").all():
        ws.append([
            s.student_id,
            s.name,
            s.age,
            s.major,
            str(s.class_info) if s.class_info else "",
            str(s.enrollment_date),
            str(s.graduation_date),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
    wb.save(response)
    return response
